"""
refresh_data.py
---------------
Template script to re-download and re-clean RBI DBIE data.

HOW TO USE:
1. Go to each DBIE URL below, download the latest XLSX manually
2. Place downloaded files in a /raw folder next to this script
3. Run: python scripts/refresh_data.py
4. Cleaned CSVs will be updated in /data

DBIE SOURCE URLs (open in browser, click Download → Excel):
- Select Aggregates (CASA):
    https://data.rbi.org.in/DBIE/dbie.rbi?site=publications#!20
- Gross & Net NPA Bank Group:
    https://data.rbi.org.in/DBIE/dbie.rbi?site=publications#!4
- Payment System Indicators:
    https://data.rbi.org.in/DBIE/dbie.rbi?site=publications#!4
- Quarterly Spatial Distribution (State CD Ratio):
    https://data.rbi.org.in → Publications → Quarterly Spatial Distribution → Statement 3A
- Handbook (GDP/CPI/Repo):
    https://data.rbi.org.in → Publications → Handbook of Statistics → Part I Annual Series
"""

import os
import pandas as pd
import openpyxl
import csv
from datetime import datetime

RAW_DIR  = os.path.join(os.path.dirname(__file__), '..', 'raw')
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
os.makedirs(RAW_DIR,  exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)


def clean_select_aggregates(raw_path):
    """CASA ratio from Scheduled Commercial Banks - Select Aggregates."""
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for row in rows[1:]:
        if not row[0]: continue
        d = row[0]
        date_str = d.strftime("%Y-%m-%d") if hasattr(d,'strftime') else str(d)
        demand, time_dep, agg_dep, bank_credit, casa = row[1], row[2], row[3], row[14], row[19]
        out.append([date_str, demand, time_dep, agg_dep, bank_credit, casa])
    out_path = os.path.join(DATA_DIR, 'clean_casa_aggregates.csv')
    with open(out_path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(["Date","Demand_Deposits_Cr","Time_Deposits_Cr",
                    "Aggregate_Deposits_Cr","Bank_Credit_Cr","CASA_Ratio"])
        w.writerows(out)
    print(f"  ✓ CASA aggregates → {len(out)} rows")


def clean_npa_bankgroup(raw_path):
    """Gross & Net NPA by bank group."""
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb['Report 1']
    rows = list(ws.iter_rows(values_only=True))
    sections = {6:"Scheduled Commercial Banks (All)", 42:"Public Sector Banks",
                78:"Old Private Sector Banks", 105:"Private Sector Banks",
                141:"Foreign Banks In India", 177:"Small Finance Banks"}
    import re
    out = []
    current_section = None
    for i, row in enumerate(rows, start=1):
        if i in sections:
            current_section = sections[i]
            continue
        if not current_section: continue
        year_cell = row[1]
        if isinstance(year_cell, str) and re.match(r'^\d{4}-\d{2}', year_cell.strip()):
            try:
                out.append([current_section, year_cell.strip()[:7],
                            row[2], row[3], row[4], row[5], row[7], row[8]])
            except: continue
    out_path = os.path.join(DATA_DIR, 'clean_npa_bankgroup.csv')
    with open(out_path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(["Bank_Group","Year","Gross_Advances","Net_Advances",
                    "Gross_NPA_Amount","Gross_NPA_Pct","Net_NPA_Amount","Net_NPA_Pct"])
        w.writerows(out)
    print(f"  ✓ NPA bank group → {len(out)} rows")


def clean_state_cd_ratio(raw_path):
    """State-wise CD ratio from Quarterly Spatial Distribution 3A."""
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    row7 = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
    quarters_ff, last = [], None
    for v in row6:
        if v and isinstance(v,str) and 'Q' in v: last = v.strip()
        quarters_ff.append(last)
    col_labels = {i: row7[i] for i in range(len(row7))}
    qtr_pairs = {}
    for i, (q, label) in enumerate(zip(quarters_ff, row7)):
        if not q or not label: continue
        label = str(label)
        if q not in qtr_pairs: qtr_pairs[q] = {}
        if 'Deposit' in label: qtr_pairs[q]['dep'] = i
        elif 'Credit' in label: qtr_pairs[q]['cred'] = i
    out = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        state = row[3]
        if not state or not isinstance(state, str): continue
        state = state.strip()
        if not state: continue
        for qtr, cols in qtr_pairs.items():
            if 'dep' not in cols or 'cred' not in cols: continue
            dep, cred = row[cols['dep']], row[cols['cred']]
            if dep is None or cred is None: continue
            try: out.append([state, str(row[2] or '').strip(), qtr, float(dep), float(cred)])
            except: continue
    df = pd.DataFrame(out, columns=["State","Region","Quarter","Deposits_Cr","Credit_Cr"])
    state_agg = df.groupby(["State","Quarter"])[["Deposits_Cr","Credit_Cr"]].sum().reset_index()
    state_agg["CD_Ratio"] = (state_agg["Credit_Cr"]/state_agg["Deposits_Cr"]*100).round(1)
    q4 = state_agg[state_agg["Quarter"].str.contains("Q4")].copy()
    q4["FY"] = q4["Quarter"].str[:7]
    q4 = q4.drop(columns=["Quarter"]).sort_values(["FY","CD_Ratio"], ascending=[True,False])
    q4.to_csv(os.path.join(DATA_DIR, 'clean_state_cd_ratio.csv'), index=False)
    print(f"  ✓ State CD ratio → {len(q4)} rows")


# ── Run all cleaners if raw files exist ──────────────────────────────────────
CLEANERS = {
    "Scheduled_Commercial_Banks_-_Select_Aggregates.xlsx": clean_select_aggregates,
    "Gross_and_Net_NPAs_of_Scheduled_Commercial_Banks_-_Bank_Group-Wise.xlsx": clean_npa_bankgroup,
    "Statement_No__3A__State-wise.xlsx": clean_state_cd_ratio,
}

print("Checking for raw files in /raw ...")
ran = 0
for filename, cleaner in CLEANERS.items():
    path = os.path.join(RAW_DIR, filename)
    if os.path.exists(path):
        print(f"Processing {filename} ...")
        cleaner(path)
        ran += 1
    else:
        print(f"  ⚠ Not found: {filename} — skipping")

if ran == 0:
    print("\nNo raw files found. Download from DBIE (see URLs at top of this file) "
          "and place in /raw before running.")
else:
    print(f"\nDone — {ran} file(s) refreshed.")
